# Standard library imports
import logging
from decimal import Decimal

# Third-party imports
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

# Django imports
from django.conf import settings as django_settings
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.db import IntegrityError, transaction
from django.forms import ValidationError
from django.http import HttpResponse, HttpResponseServerError
from django.shortcuts import render, redirect, get_object_or_404
from django.views import View

# Local application/library specific imports
from accounts.forms import UserTimeZoneForm
from accounts.models import Learner, Supervisor
from administrator.forms import AdminNameForm, AdminEmailForm, AdminProfilePictureForm
from courses.forms import CourseDeliveryForm, ScormCloudCourseForm
from courses.models import Attendance, CourseDelivery, Enrollment, Feedback, ScormCloudCourse, ScormCloudRegistration
from learner.forms import LearnerForm
from supervisor.forms import SupervisorForm

logger = logging.getLogger(__name__)

User = get_user_model()

@login_required
def dashboard(request):
    print(django_settings)
    print(django_settings.DOMAIN_NAME)
    try:
        return render(request, 'administrator/dashboard.html')
    except Exception as e:
        logger.error(f"Error loading dashboard: {e}")
        return HttpResponseServerError("An error occurred")
    
@login_required
def calendar(request):
    try:
        return render(request, 'administrator/calendar.html')
    except Exception as e:
        logger.error(f"Error loading calendar: {e}")
        return HttpResponseServerError("An error occurred")

@login_required
def leaderboard(request):
    try:
        return render(request, 'administrator/leaderboard.html')
    except Exception as e:
        logger.error(f"Error loading leaderboard: {e}")
        return HttpResponseServerError("An error occurred")

@login_required
def course_list(request):
    try:
        courses = ScormCloudCourse.objects.all()
    except ScormCloudCourse.DoesNotExist:
        courses = None
    return render(request, 'administrator/course_list.html', {'courses': courses})

@login_required
def upload_course(request):
    domain = django_settings.DOMAIN_NAME
    if request.method == 'POST':
        form = ScormCloudCourseForm(request.POST, request.FILES)
        if form.is_valid():
            course_id = form.cleaned_data['course_id']
            file = form.cleaned_data['file']

            category = form.cleaned_data.get('category', '')
            duration = form.cleaned_data.get('duration', '')
            short_description = form.cleaned_data.get('short_description', '')
            long_description = form.cleaned_data.get('long_description', '')
            cover_image = form.cleaned_data.get('cover_image', '')

            files = {'file': file}

            try:
                response = requests.post(
                    f'{domain}/api/create_course/{course_id}/',
                    files=files,
                    timeout=300,
                )

                response.raise_for_status()
                api_response_data = response.json()

                if api_response_data.get('message') == 'Course created successfully':
                    course, created = ScormCloudCourse.objects.get_or_create(
                        course_id=course_id, 
                        defaults={
                            'title': form.cleaned_data['title'],
                            'version': api_response_data.get('version', ''),
                            'web_path': api_response_data.get('web_path_to_course', '')  
                        }
                    )
                    if not created:
                        course.category = category
                        course.duration = duration
                        course.short_description = short_description
                        course.long_description = long_description
                        course.cover_image = cover_image
                        course.save()

                    logger.info(f"Course created/updated successfully: {course_id}, Response: {api_response_data}")

                    # Register the administrator as a learner for this course
                    try:
                        admin_user = request.user
                        learner = Learner.objects.get(user=admin_user)
                        
                        register_url = f'{domain}/api/register/'
                        register_data = {
                            'course_id': course_id,
                            'learner_id': learner.id
                        }
                        
                        register_response = requests.post(register_url, data=register_data)
                        register_response.raise_for_status()
                        
                        logger.info(f"Administrator registered for course: {course_id}, Learner ID: {learner.id}")
                    except Learner.DoesNotExist:
                        logger.error(f"Learner object not found for administrator: {admin_user.username}")
                    except requests.RequestException as e:
                        logger.error(f"Error registering administrator for course: {e}")

                    return redirect('course_list')

                else:
                    error_msg = api_response_data.get('error', 'Unknown error')
                    logger.error(f"Course creation API failed: {error_msg}")
                    form.add_error(None, error_msg)

            except requests.RequestException as e:
                logger.exception("Error calling create_course API:")
                form.add_error(None, f"Error creating course: {e}")
            except ScormCloudCourse.DoesNotExist:
                logger.error(f"Course with ID {course_id} not found after API call.")
                form.add_error(None, "Course not found in database after creation.")
    else:
        form = ScormCloudCourseForm()
    return render(request, 'administrator/upload_course.html', {
        'form': form, 
        'generated_course_id': form.fields['course_id'].initial
        })


@login_required
def administrator_settings(request):
    forms = {
        'timezone_form': UserTimeZoneForm(prefix='timezone_form', instance=request.user),
        'admin_name_form': AdminNameForm(prefix='admin_name_form', instance=request.user),
        'admin_email_form': AdminEmailForm(prefix='admin_email_form', instance=request.user),
        'profile_picture_form': AdminProfilePictureForm(prefix='profile_picture_form', instance=request.user),
    }
    try:
        if request.method == 'POST':
            for form_name, form in forms.items():
                if form_name in request.POST:
                    if form_name == 'profile_picture_form':
                        form = form.__class__(request.POST, request.FILES, instance=request.user, prefix=form_name)
                    else:
                        form = form.__class__(request.POST, instance=request.user, prefix=form_name)
                    if form.is_valid():
                        form.save()
                        return redirect('administrator_settings')
                    else:
                        logger.error(f"Form errors in {form_name}: {form.errors.as_json()}")
                        forms[form_name] = form
                        break  
        return render(request, 'administrator/settings.html', {'forms': forms})
    except Exception as e:
        logger.error(f"Error loading profile: {e}")
        return HttpResponseServerError("An error occurred")
    
def learner_list(request):
    try:
        learners = Learner.objects.all()
    except Learner.DoesNotExist:
        learners = None
    return render(request, 'administrator/learner_list.html', {'learners': learners})

def supervisor_list(request):
    try:
        supervisors = Supervisor.objects.all()
    except Supervisor.DoesNotExist:
        supervisors = None
    return render(request, 'administrator/supervisor_list.html', {'supervisors': supervisors})


@transaction.atomic
def register_learner(request):
    if request.method == 'POST':
        form = LearnerForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    form.save()
                messages.success(request, 'Learner registered successfully!')
                return redirect('administrator_learner_list')
            except IntegrityError:
                messages.error(request, 'A learner with this email already exists.')
                # Explicitly return to avoid further operations in the same transaction block
                return render(request, 'administrator/register_learner.html', {'form': form})
            except ValidationError as e:
                for field, errors in e.message_dict.items():
                    for error in errors:
                        messages.error(request, f'{field.capitalize()}: {error}')
                # Return here if you want to stop further processing and show the form again
                return render(request, 'administrator/register_learner.html', {'form': form})
            except Exception as e:
                logger.error(f'Error registering learner: {e}')
                messages.error(request, 'An error occurred during registration. Please try again.')
                # Return here to handle any other exceptions and stop further processing
                return render(request, 'administrator/register_learner.html', {'form': form})
        else:
            # Iterate through form errors and add them as messages
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field.capitalize()}: {error}')
    else:
        form = LearnerForm()

    return render(request, 'administrator/register_learner.html', {'form': form})

@transaction.atomic
def register_supervisor(request):
    if request.method == 'POST':
        form = SupervisorForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    form.save()
                messages.success(request, 'Supervisor registered successfully!')
                return redirect('administrator_supervisor_list')
            except IntegrityError:
                messages.error(request, 'A supervisor with this email already exists.')
                # Explicitly return to avoid further operations in the same transaction block
                return render(request, 'administrator/register_supervisor.html', {'form': form})
            except ValidationError as e:
                for field, errors in e.message_dict.items():
                    for error in errors:
                        messages.error(request, f'{field.capitalize()}: {error}')
                # Return here if you want to stop further processing and show the form again
                return render(request, 'administrator/register_supervisor.html', {'form': form})
            except Exception as e:
                logger.error(f'Error registering supervisor: {e}')
                messages.error(request, 'An error occurred during registration. Please try again.')
                # Return here to handle any other exceptions and stop further processing
                return render(request, 'administrator/register_supervisor.html', {'form': form})
        else:
            # Iterate through form errors and add them as messages
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field.capitalize()}: {error}')
    else:
        form = SupervisorForm()

    return render(request, 'administrator/register_supervisor.html', {'form': form})


def enrollment_activity(request):
    try:
        return render(request, 'administrator/enrollment_activity.html')
    except Exception as e:
        logger.error(f"Error loading enrollment_activity: {e}")
        return HttpResponseServerError("An error occurred")
    
def progress_performance(request):
    try:
        return render(request, 'administrator/progress_performance.html')
    except Exception as e:
        logger.error(f"Error loading progress_performance: {e}")
        return HttpResponseServerError("An error occurred")
    
def engagement_feedback(request):
    try:
        return render(request, 'administrator/engagement_feedback.html')
    except Exception as e:
        logger.error(f"Error loading engagement_feedback: {e}")
        return HttpResponseServerError("An error occurred")

def system_usuage(request):
    try:
        return render(request, 'administrator/system_usuage.html')
    except Exception as e:
        logger.error(f"Error loading system_usuage: {e}")
        return HttpResponseServerError("An error occurred")
    

def course_delivery_list(request, pk):
    try:
        course = ScormCloudCourse.objects.get(pk=pk)
        deliveries = course.deliveries.all()
    except ScormCloudCourse.DoesNotExist:
        course = None
        deliveries = None
    return render(request, 'administrator/course_delivery/course_delivery_list.html', {'course': course, 'deliveries': deliveries})

def course_delivery_detail(request, course_id, delivery_id):
    try:
        course = ScormCloudCourse.objects.get(id=course_id)
        delivery = CourseDelivery.objects.get(course=course, delivery_code=delivery_id)
    except ScormCloudCourse.DoesNotExist:
        course = None
        delivery = None
    return render(request, 'administrator/course_delivery/course_delivery_detail.html', {'course': course, 'delivery': delivery})

class CourseDeliveryCreateView(View):
    def get(self, request, pk=None):
        logger.debug("Entering GET method")
        try:
            form = CourseDeliveryForm()
            course = None
            if pk:
                course = get_object_or_404(ScormCloudCourse, pk=pk)
                logger.info(f"Course with pk={pk} found")
            return render(request, 'administrator/course_delivery/course_delivery_create.html', {'form': form, 'course': course})
        except Exception as e:
            logger.error(f"Error in GET method: {e}")
            # Handle the error or redirect to an error page
            return render(request, 'errors/error_page.html', {'error': e})

    def post(self, request, pk=None):
        logger.debug("Entering POST method")
        print(f"Entering POST method with pk={pk} and data={request.POST}")
        try:
            form = CourseDeliveryForm(request.POST)
            if form.is_valid():
                course_delivery = form.save(commit=False)
                # Fetch the course using the pk and assign it to the course_delivery object
                if pk:
                    course = get_object_or_404(ScormCloudCourse, pk=pk)
                    course_delivery.course = course  
                    course_delivery.status = 'INACTIVE'
                    course_delivery.created_by = request.user
                    course_delivery.timezone = request.user.timezone
                    logger.info(f"Assigned course with pk={pk} to course delivery")
                    print(f"Assigned course with pk={pk} to course delivery")
                
                if not course_delivery.delivery_code:
                    course_delivery.delivery_code = CourseDelivery.generate_unique_delivery_code()
                    logger.info("Generated unique delivery code")
                    print(f"Generated unique delivery code: {course_delivery.delivery_code}")
                
                # Set the course title based on the delivery type
                delivery_type = form.cleaned_data.get('delivery_type')
                if delivery_type == 'SELF_PACED':
                    course_delivery.title = f"{course.title} - Self Paced"
                elif delivery_type == 'INSTRUCTOR_LED':
                    course_delivery.title = f"{course.title} - Instructor Led"
                logger.info(f"Course title set based on delivery type: {course_delivery.title}")
                
                course_delivery.save()
                form.save_m2m()
                logger.info("Course delivery saved successfully")
                print("Course delivery saved successfully")
                return redirect('course_delivery_list', pk=course.pk)
            else:
                logger.warning(f"Form is not valid: {form.errors}")
            
            course = None
            if pk:
                course = get_object_or_404(ScormCloudCourse, pk=pk)
                logger.info(f"Course with pk={pk} found for POST")
                print(f"Course with pk={pk} found for POST")
            
            return render(request, 'administrator/course_delivery/course_delivery_create.html', {'form': form, 'course': course})
        except Exception as e:
            logger.error(f"Error in POST method: {e}")
            print(f"Error in POST method: {e}")
            # Handle the error or redirect to an error page
            return render(request, 'errors/error_page.html', {'error': e})
        

def export_attendance(request, delivery_id):
    delivery = get_object_or_404(CourseDelivery, id=delivery_id)
    participants = Learner.objects.filter(enrolled_deliveries=delivery)

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Add course delivery information at the top
    course_info = [
        ['Course Title', delivery.course.title],
        ['Delivery Code', delivery.delivery_code],
        ['Start Date', delivery.start_date.strftime('%Y-%m-%d')],
        ['End Date', delivery.end_date.strftime('%Y-%m-%d')],
        ['Deactivation Date', delivery.deactivation_date.strftime('%Y-%m-%d')],
    ]

    for row in course_info:
        ws.append(row)
        ws.cell(ws.max_row, 1).font = Font(bold=True)
        ws.cell(ws.max_row, 2).alignment = Alignment(horizontal='left')

    # Add a blank row
    ws.append([])

    # Add participant details header
    participant_headers = ['Participant Name', 'Email', 'Total Time Spent (hours)']
    ws.append(participant_headers)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    # Add participant data
    for participant in participants:
        ws.append([
            participant.user.first_name,
            participant.user.email,
            # participant.get_total_time_spent()  # You'll need to implement this method
        ])

    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Create the HttpResponse object with Excel mime type
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Course_Attendance_{delivery.delivery_code}.xlsx'

    # Save the workbook to the response
    wb.save(response)

    return response


def preview_course(request, learner_id, course_id):
    try:
        learner = get_object_or_404(Learner, id=learner_id)
        course = get_object_or_404(ScormCloudCourse, id=course_id)
        registration = get_object_or_404(ScormCloudRegistration, learner=learner, course_id=course.course_id)
        return render(request, 'administrator/preview_course.html', { 
            'learner_id': learner_id,
            'registration_id': registration.registration_id,
            'course': course,
        })
    except Exception as e:
        logger.error(f"Error loading preview_course: {e}")
        return HttpResponseServerError("An error occurred")


def support(request):
    try:
        return render(request, 'administrator/support.html')
    except Exception as e:
        logger.error(f"Error loading support: {e}")
        return HttpResponseServerError("An error occurred")